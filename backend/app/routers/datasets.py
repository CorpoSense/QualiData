"""Dataset routes for import/export."""

import tempfile
import os
import json
import logging

from app.services.smart_importer import SmartImporter

import io

import pandas as pd
import numpy as np
from datetime import datetime, date, time, timedelta
from fastapi import (
    APIRouter,
    Depends,
    File,
    Form,
    HTTPException,
    Query,
    UploadFile,
    status,
)
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.database import get_async_session
from app.db.models import Dataset, Project, User
from app.routers.auth import get_current_active_user

# For Excel sheet detection
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

router = APIRouter(prefix="/datasets", tags=["datasets"])

logger = logging.getLogger(__name__)


# Pydantic schemas
class DatasetCreate(BaseModel):
    name: str
    description: str | None = None
    project_id: str


class DatasetUpdate(BaseModel):
    name: str | None = None
    description: str | None = None


class DatasetResponse(BaseModel):
    id: str
    name: str
    description: str | None
    project_id: str
    file_name: str | None
    file_size: int
    file_type: str | None
    row_count: int
    columns: list | None
    # preview_data: list | None
    created_at: datetime | None

    model_config = {"from_attributes": True}


class MultiImportResult(BaseModel):
    file_name: str
    success: bool
    message: str | None
    dataset_id: str | None
    row_count: int


class MultiImportResponse(BaseModel):
    status: str
    message: str
    results: list[MultiImportResult]
    total_rows: int
    total_size: int


class ColumnInfo(BaseModel):
    name: str
    dtype: str


class DatasetPreviewResponse(BaseModel):
    columns: list[ColumnInfo]
    preview_data: list[dict]
    row_count: int


def detect_columns(df: pd.DataFrame) -> list[dict]:
    """Detect column names and types from DataFrame."""
    columns = []
    for col in df.columns:
        dtype = str(df[col].dtype)
        # Map pandas dtypes to simpler types
        if "int" in dtype:
            dtype = "integer"
        elif "float" in dtype:
            dtype = "float"
        elif "object" in dtype:
            dtype = "string"
        elif "datetime" in dtype:
            dtype = "datetime"
        elif "bool" in dtype:
            dtype = "boolean"
        columns.append({"name": col, "dtype": dtype})
    return columns


def _normalize_columns(columns):
    """Normalize columns to list of dicts, handling legacy string format."""
    if not columns:
        return []
    if columns and isinstance(columns[0], str):
        return [{"name": c, "dtype": "string"} for c in columns]
    return columns


def _ensure_json_serializable(data: list[dict]) -> list[dict]:
    """Ensure all values in the data are JSON serializable."""
    def convert_value(val):
        if val is None:
            return None
        if isinstance(val, (datetime, date, time)):
            return val.isoformat()
        if isinstance(val, timedelta):
            return str(val)
        if isinstance(val, (np.integer,)):
            return int(val)
        if isinstance(val, (np.floating,)):
            return float(val) if not np.isnan(val) else None
        if isinstance(val, np.ndarray):
            return val.tolist()
        if isinstance(val, pd.Timestamp):
            return val.isoformat()
        if isinstance(val, pd.Timedelta):
            return str(val)
        # Try to convert to string as fallback
        try:
            json.dumps(val)
            return val
        except (TypeError, ValueError):
            return str(val)
    
    return [{k: convert_value(v) for k, v in row.items()} for row in data]


def get_preview_data(df: pd.DataFrame, max_rows: int = 500, offset: int = 0) -> list[dict]:
    """Get preview data from DataFrame with proper JSON serialization."""
    # Create a copy to avoid modifying original
    df_copy = df.copy()
    
    # Convert all datetime columns to ISO format strings
    for col in df_copy.columns:
        if pd.api.types.is_datetime64_any_dtype(df_copy[col]):
            df_copy[col] = df_copy[col].astype(str)
            df_copy[col] = df_copy[col].replace("NaT", None)
        # Handle timedelta
        elif pd.api.types.is_timedelta64_dtype(df_copy[col]):
            df_copy[col] = df_copy[col].astype(str)
        # Handle period
        # elif pd.api.types.is_period_dtype(df_copy[col]):
        elif isinstance(df_copy[col], pd.PeriodDtype):
            df_copy[col] = df_copy[col].astype(str)
    
    # Replace NaN/inf with None for JSON compatibility
    df_copy = df_copy.replace([np.inf, -np.inf], np.nan).replace({np.nan: None})
    
    # Convert to dict and ensure all values are JSON serializable
    records = df_copy.iloc[offset:offset + max_rows].to_dict(orient="records")
    
    # Final pass to ensure JSON serialization
    return _ensure_json_serializable(records)


def get_full_data_json(df: pd.DataFrame) -> dict:
    """Convert full DataFrame to JSON-serializable format for storage in data_json field."""
    # Create a copy to avoid modifying original
    df_copy = df.copy()
    
    # Convert all datetime columns to ISO format strings
    for col in df_copy.columns:
        if pd.api.types.is_datetime64_any_dtype(df_copy[col]):
            df_copy[col] = df_copy[col].astype(str)
            df_copy[col] = df_copy[col].replace("NaT", None)
        # Handle timedelta
        elif pd.api.types.is_timedelta64_dtype(df_copy[col]):
            df_copy[col] = df_copy[col].astype(str)
        # Handle period
        # elif pd.api.types.is_period_dtype(df_copy[col]):
        elif isinstance(df_copy[col], pd.PeriodDtype):
        # elif isinstance(df_copy[col], pd.PeriodDtype, pd.IntervalDtype):
            df_copy[col] = df_copy[col].astype(str)
    
    # Replace NaN/inf with None for JSON compatibility
    df_copy = df_copy.replace([np.inf, -np.inf], np.nan).replace({np.nan: None})
    
    # Convert to dict and ensure all values are JSON serializable
    records = df_copy.to_dict(orient="records")
    
    # Final pass to ensure JSON serialization
    return {"data": _ensure_json_serializable(records)}


# Routes
@router.post("/import/single", response_model=DatasetResponse)
async def import_single_dataset(
    file: UploadFile = File(...),
    project_id: str = Form(...),
    name: str | None = Form(None),
    description: str | None = Form(None),
    # Smart import options
    auto_detect: str | None = Form('true'),
    delimiter: str | None = Form(None),
    encoding: str | None = Form(None),
    has_header: str | None = Form(None),
    sheet_name: str | None = Form(None),
    current_user: User = Depends(get_current_active_user),
    session: AsyncSession = Depends(get_async_session),
):
    """Import a single data file (CSV, Excel, or JSON)."""
    # Verify project belongs to user
    result = await session.execute(
        select(Project).where(
            Project.id == project_id, Project.user_id == current_user.id
        )
    )
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Project not found"
        )

    # Save uploaded file to temp file for smart importer
    filename = file.filename or f"data_{int(time.time())}.csv"
    content = await file.read()
    file_size = len(content)
    
    # Determine file type
    file_ext = os.path.splitext(filename)[1].lower()
    file_type = "csv"
    if file_ext in (".xlsx", ".xls", ".xlsb", ".ods"):
        file_type = "excel"
    elif file_ext == ".json":
        file_type = "json"
    elif file_ext == ".parquet":
        file_type = "parquet"
    
    # Convert string values from FormData
    auto_detect_bool = auto_detect.lower() == 'true' if auto_detect else True
    has_header_bool = has_header.lower() == 'true' if has_header else True
    
    # Save to temp file
    with tempfile.NamedTemporaryFile(delete=False, suffix=file_ext) as tmp:
        tmp.write(content)
        tmp_path = tmp.name
    
    try:
        # Parquet files are binary and don't need smart importer or encoding detection
        if file_type == "parquet":
            df = pd.read_parquet(tmp_path)
        elif auto_detect:
            # Use smart importer for automatic detection
            importer = SmartImporter()
            analysis = importer.analyze(tmp_path)
            
            if not analysis.is_importable:
                error_msgs = [m.message for m in analysis.messages if m.severity.value == "error"]
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Cannot analyze file: {'; '.join(error_msgs)}"
                )
            
            df = importer.import_file(
                tmp_path,
                analysis=analysis,
                delimiter=delimiter if not auto_detect_bool else None,
                encoding=encoding,
                has_header=has_header_bool if not auto_detect_bool else None,
                sheet_name=sheet_name,
            )
        else:
            # Use manual settings (no auto detection)
            if file_type == "csv":
                df = pd.read_csv(
                    tmp_path,
                    sep=delimiter or ',',
                    encoding=encoding or 'utf-8',
                    header=0 if has_header else None,
                )
            elif file_type == "excel":
                df = pd.read_excel(
                    tmp_path,
                    sheet_name=sheet_name,
                    header=0 if has_header else None,
                )
            elif file_type == "json":
                df = pd.read_json(tmp_path)
            elif file_type == "parquet":
                df = pd.read_parquet(tmp_path)
            else:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Unsupported file type.",
                )
    finally:
        # Clean up temp file
        os.unlink(tmp_path)

    # Get column info and preview
    try:
        columns = detect_columns(df)
        preview_data = get_preview_data(df)
        data_json = get_full_data_json(df)
        row_count = len(df)
    except Exception as e:
        logger.exception("Failed to process DataFrame for import")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to process data: {str(e)}"
        )

    # Use filename as name if not provided
    if not name:
        name = filename

    # Create dataset record
    try:
        dataset = Dataset(
            name=name,
            description=description,
            project_id=project_id,
            file_name=filename,
            file_size=file_size,
            file_type=file_type,
            columns=columns,
            data_json=data_json,
            row_count=row_count,
        )
        session.add(dataset)

        # Update project stats
        project.row_count += row_count
        project.storage_bytes += file_size

        await session.commit()
        await session.refresh(dataset)

        return dataset
    except Exception as e:
        await session.rollback()
        logger.exception("Failed to save dataset to database")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to save dataset: {str(e)}"
        )


@router.post("/import/excel/sheets")
async def get_excel_sheets(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_active_user),
):
    """Get sheet names from an Excel file."""
    filename = file.filename or "data.xlsx"
    file_ext = os.path.splitext(filename)[1].lower()
    
    # Only process Excel files
    if file_ext not in (".xlsx", ".xls", ".xlsb", ".ods"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File must be an Excel file (.xlsx, .xls, .xlsb, .ods)"
        )
    
    content = await file.read()
    
    # Save to temp file
    with tempfile.NamedTemporaryFile(delete=False, suffix=file_ext) as tmp:
        tmp.write(content)
        tmp_path = tmp.name
    
    try:
        sheet_names = []
        
        if file_ext == ".xlsx" and OPENPYXL_AVAILABLE:
            # Use openpyxl for .xlsx files (faster and more reliable)
            workbook = load_workbook(tmp_path, read_only=True, data_only=True)
            sheet_names = workbook.sheetnames
            workbook.close()
        else:
            # Use pandas for other formats (.xls, .xlsb, .ods)
            excel_file = pd.ExcelFile(tmp_path)
            sheet_names = excel_file.sheet_names
            excel_file.close()
        
        return {
            "status": "success",
            "sheets": sheet_names,
            "file_name": filename
        }
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to read Excel file: {str(e)}"
        )
    finally:
        # Clean up temp file
        os.unlink(tmp_path)


@router.post("/import/multiple", response_model=MultiImportResponse)
async def import_multiple_datasets(
    files: list[UploadFile] = File(...),
    project_id: str = Form(...),
    name: str | None = Form(None),
    description: str | None = Form(None),
    auto_detect: str | None = Form('true'),
    delimiter: str | None = Form(None),
    encoding: str | None = Form(None),
    has_header: str | None = Form(None),
    sheet_name: str | None = Form(None),
    merge_strategy: str | None = Form('union'),
    current_user: User = Depends(get_current_active_user),
    session: AsyncSession = Depends(get_async_session),
):
    """Import multiple data files at once."""
    # Verify project belongs to user
    result = await session.execute(
        select(Project).where(
            Project.id == project_id, Project.user_id == current_user.id
        )
    )
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Project not found"
        )

    results = []
    total_rows = 0
    total_size = 0
    dataframes = []  # For merging when strategy is provided

    for file in files:
        file_result = {
            "file_name": file.filename or "unknown",
            "success": False,
            "message": None,
            "dataset_id": None,
            "row_count": 0
        }
        
        try:
            # Process each file (similar to existing import logic)
            filename = file.filename or "data.csv"
            content = await file.read()
            file_size = len(content)
            
            # Determine file type
            file_ext = os.path.splitext(filename)[1].lower()
            file_type = "csv"
            if file_ext in (".xlsx", ".xls", ".xlsb", ".ods"):
                file_type = "excel"
            elif file_ext == ".json":
                file_type = "json"
            elif file_ext == ".parquet":
                file_type = "parquet"
            
            # Convert string values from FormData
            auto_detect_bool = auto_detect.lower() == 'true' if auto_detect else True
            has_header_bool = has_header.lower() == 'true' if has_header else True
            
            # Save to temp file
            with tempfile.NamedTemporaryFile(delete=False, suffix=file_ext) as tmp:
                tmp.write(content)
                tmp_path = tmp.name
            
            try:
                # Parquet files are binary and don't need smart importer or encoding detection
                if file_type == "parquet":
                    df = pd.read_parquet(tmp_path)
                elif auto_detect_bool:
                    # Use smart importer
                    importer = SmartImporter()
                    analysis = importer.analyze(tmp_path)
                    
                    if not analysis.is_importable:
                        error_msgs = [m.message for m in analysis.messages if m.severity.value == "error"]
                        file_result["message"] = f"Cannot analyze file: {'; '.join(error_msgs)}"
                        results.append(file_result)
                        continue
                    
                    df = importer.import_file(
                        tmp_path,
                        analysis=analysis,
                        delimiter=delimiter if not auto_detect_bool else None,
                        encoding=encoding,
                        has_header=has_header_bool if not auto_detect_bool else None,
                        sheet_name=sheet_name,
                    )
                else:
                    # Manual settings
                    if file_type == "csv":
                        df = pd.read_csv(
                            tmp_path,
                            sep=delimiter or ',',
                            encoding=encoding or 'utf-8',
                            header=0 if has_header_bool else None,
                        )
                    elif file_type == "excel":
                        df = pd.read_excel(
                            tmp_path,
                            sheet_name=sheet_name,
                            header=0 if has_header_bool else None,
                        )
                    elif file_type == "json":
                        df = pd.read_json(tmp_path)
                    elif file_type == "parquet":
                        df = pd.read_parquet(tmp_path)
                    else:
                        file_result["message"] = "Unsupported file type"
                        results.append(file_result)
                        continue
            finally:
                os.unlink(tmp_path)
            
            # Store DataFrame for potential merging
            dataframes.append({
                "df": df,
                "filename": filename,
                "file_size": file_size,
                "file_type": file_type,
                "file_result": file_result
            })
            
        except Exception as e:
            file_result["message"] = str(e)
            results.append(file_result)
    
    # Handle merging if strategy is provided and multiple files exist
    if merge_strategy and len(dataframes) > 1:
        try:
            # Determine final columns based on strategy
            all_columns = set()
            for item in dataframes:
                all_columns.update(item["df"].columns)
            
            if merge_strategy == "intersection":
                common_cols = set(dataframes[0]["df"].columns)
                for item in dataframes[1:]:
                    common_cols &= set(item["df"].columns)
                final_columns = sorted(common_cols)
            elif merge_strategy == "strict":
                ref_cols = set(dataframes[0]["df"].columns)
                mismatches = []
                for i, item in enumerate(dataframes[1:], 2):
                    ds_cols = set(item["df"].columns)
                    if ds_cols != ref_cols:
                        missing = ref_cols - ds_cols
                        extra = ds_cols - ref_cols
                        detail = f"File {i} ({item['filename']})"
                        if missing:
                            detail += f" is missing: {sorted(missing)}"
                        if extra:
                            detail += f" has extra: {sorted(extra)}"
                        mismatches.append(detail)
                if mismatches:
                    return {
                        "status": "failed",
                        "message": f"Column mismatch in strict mode: {'; '.join(mismatches)}",
                        "results": [],
                        "total_rows": 0,
                        "total_size": 0
                    }
                final_columns = sorted(ref_cols)
            else:  # union
                final_columns = sorted(all_columns)
            
            if not final_columns:
                raise HTTPException(status_code=400, detail="No columns to merge")
            
            # Align all DataFrames to final columns
            aligned_dfs = []
            for item in dataframes:
                df = item["df"]
                for col in final_columns:
                    if col not in df.columns:
                        df[col] = None
                aligned_dfs.append(df[final_columns])
            
            # Concatenate
            merged_df = pd.concat(aligned_dfs, ignore_index=True)
            
            # Get column info and preview
            columns = detect_columns(merged_df)
            preview_data = get_preview_data(merged_df)
            data_json = get_full_data_json(merged_df)
            row_count = len(merged_df)
            
            # Calculate total file size
            total_file_size = sum(item["file_size"] for item in dataframes)
            
            # Use provided name or default
            dataset_name = name if name else "Merged Dataset"
            
            # Create merged dataset record
            dataset = Dataset(
                name=dataset_name,
                description=description or f"Merged from {len(dataframes)} files ({merge_strategy})",
                project_id=project_id,
                file_name=f"{dataset_name}.csv",
                file_size=total_file_size,
                file_type="csv",
                columns=columns,
                data_json=data_json,
                row_count=row_count,
            )
            session.add(dataset)
            await session.flush()  # Get the ID
            
            # Create success results for all files
            for item in dataframes:
                item["file_result"]["success"] = True
                item["file_result"]["dataset_id"] = dataset.id
                item["file_result"]["row_count"] = len(item["df"])
                item["file_result"]["message"] = f"Merged into '{dataset_name}'"
                results.append(item["file_result"])
            
            total_rows = row_count
            total_size = total_file_size
            
        except Exception as e:
            logger.exception("Failed to merge datasets")
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Failed to merge datasets: {str(e)}"
            )
    else:
        # Import each file as a separate dataset
        for item in dataframes:
            df = item["df"]
            filename = item["filename"]
            file_size = item["file_size"]
            file_type = item["file_type"]
            file_result = item["file_result"]
            
            try:
                # Get column info and preview
                columns = detect_columns(df)
                preview_data = get_preview_data(df)
                row_count = len(df)
                
                # Use filename as name if not provided
                dataset_name = name if name and len(files) == 1 else filename
                
                # Create dataset record
                dataset = Dataset(
                    name=dataset_name,
                    description=description,
                    project_id=project_id,
                    file_name=filename,
                    file_size=file_size,
                    file_type=file_type,
                    columns=columns,
                    data_json=data_json,
                    row_count=row_count,
                )
                session.add(dataset)
                await session.flush()  # Get the ID
                
                file_result["success"] = True
                file_result["dataset_id"] = dataset.id
                file_result["row_count"] = row_count
                file_result["message"] = f"Imported {row_count} rows"
                
                total_rows += row_count
                total_size += file_size
            except Exception as e:
                await session.rollback()
                logger.exception("Failed to save dataset to database")
                file_result["message"] = f"Failed to save dataset: {str(e)}"
            
            results.append(file_result)
    
    # Update project stats
    project.row_count += total_rows
    project.storage_bytes += total_size
    
    await session.commit()
    
    return {
        "status": "success",
        "message": f"Imported {len([r for r in results if r['success']])} of {len(files)} files",
        "results": results,
        "total_rows": total_rows,
        "total_size": total_size
    }


@router.get("/{dataset_id}/preview", response_model=DatasetPreviewResponse)
async def preview_dataset(
    dataset_id: str,
    limit: int = Query(10, ge=1, le=500),
    page: int = Query(1, ge=1),
    current_user: User = Depends(get_current_active_user),
    session: AsyncSession = Depends(get_async_session),
):
    """Get dataset preview (columns and sample data)."""
    result = await session.execute(select(Dataset).where(Dataset.id == dataset_id))
    dataset = result.scalar_one_or_none()

    if not dataset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Dataset not found"
        )

    # Verify ownership via project
    project_result = await session.execute(
        select(Project).where(
            Project.id == dataset.project_id, Project.user_id == current_user.id
        )
    )
    if not project_result.scalar_one_or_none():
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN, detail="Access denied"
        )

    # Use data_json for preview - generate on-the-fly
    if dataset.data_json and "data" in dataset.data_json:
        full_data = dataset.data_json["data"]
        start_idx = (page - 1) * limit
        end_idx = start_idx + limit
        preview_data = full_data[start_idx:end_idx]
        row_count = len(full_data)
    else:
        # No data available
        return {
            "columns": _normalize_columns(dataset.columns),
            "preview_data": [],
            "row_count": 0,
            "page": page,
            "limit": limit,
        }

    # Normalize columns: handle both dict format and legacy string format
    columns = _normalize_columns(dataset.columns)

    # DEBUG: Log column order in preview response to diagnose column order issue
    import logging
    _logger = logging.getLogger(__name__)
    _col_names = [c["name"] for c in columns] if columns and isinstance(columns[0], dict) else columns
    _data_keys = list(preview_data[0].keys()) if preview_data else []

    return {
        "columns": columns,
        "preview_data": preview_data,
        "row_count": row_count,
        "page": page,
        "limit": limit,
    }


@router.post("/{dataset_id}/filtered")
async def filtered_dataset_post(
    dataset_id: str,
    filters: dict,
    limit: int = Query(10, ge=1, le=500),
    page: int = Query(1, ge=1),
    current_user: User = Depends(get_current_active_user),
    session: AsyncSession = Depends(get_async_session),
):
    """Filter dataset rows by column values (POST with JSON body).

    Supports two filter formats per column:
    - Substring match (legacy): {"country": "N/A"} — case-insensitive substring match
    - Selected values (multi-filter): {"country": {"selected_values": ["France", "Germany"]}}
      — exact match where the cell value's string representation is in the selected list.
      Use null in the list to match null/empty values: {"country": {"selected_values": [null, "France"]}}

    Both formats can be mixed in the same request. All filters are AND-combined.
    Returns matching rows with pagination + total matching count.
    """
    result = await session.execute(select(Dataset).where(Dataset.id == dataset_id))
    dataset = result.scalar_one_or_none()

    if not dataset:
        raise HTTPException(status_code=404, detail="Dataset not found")

    project_result = await session.execute(
        select(Project).where(
            Project.id == dataset.project_id, Project.user_id == current_user.id
        )
    )
    if not project_result.scalar_one_or_none():
        raise HTTPException(status_code=403, detail="Access denied")

    # Use data_json for filtering - generate preview on-the-fly
    if dataset.data_json and "data" in dataset.data_json:
        preview_list = dataset.data_json["data"]
    else:
        return {"columns": _normalize_columns(dataset.columns), "preview_data": [], "row_count": 0, "total_matching": 0, "page": page, "limit": limit}

    # Parse filters into two categories:
    # 1. Substring filters: {"col": "substring"} (legacy format)
    # 2. Selected values filters: {"col": {"selected_values": ["val1", "val2"]}} (new format)
    substring_filters = {}
    selected_values_filters = {}

    for col, val in filters.items():
        if isinstance(val, dict) and "selected_values" in val:
            # New format: selected_values list
            selected_vals = val["selected_values"]
            if isinstance(selected_vals, list) and len(selected_vals) > 0:
                # Normalize: convert all non-null values to strings, keep None as-is
                normalized = set()
                for v in selected_vals:
                    if v is None:
                        normalized.add(None)
                    else:
                        normalized.add(str(v).strip())
                selected_values_filters[col] = normalized
        elif val is not None and str(val).strip():
            # Legacy format: substring match
            substring_filters[col] = str(val).strip().lower()

    # Filter rows
    matching_indices = []
    for i, row in enumerate(preview_list):
        match = True

        # Check substring filters
        for col, filter_val in substring_filters.items():
            cell_val = str(row.get(col, "")).lower()
            if filter_val not in cell_val:
                match = False
                break

        if not match:
            continue

        # Check selected_values filters (exact match on string representation)
        for col, allowed_values in selected_values_filters.items():
            cell_raw = row.get(col)
            if cell_raw is None:
                # Cell is null — check if null is in allowed values
                if None not in allowed_values:
                    match = False
                    break
            else:
                cell_str = str(cell_raw)
                if cell_str not in allowed_values:
                    match = False
                    break

        if match:
            matching_indices.append(i)

    total_matching = len(matching_indices)

    # Paginate matching indices
    start_idx = (page - 1) * limit
    end_idx = start_idx + limit
    page_indices = matching_indices[start_idx:end_idx]

    preview_data = [preview_list[i] for i in page_indices]

    return {
        "columns": _normalize_columns(dataset.columns),
        "preview_data": preview_data,
        "row_count": dataset.row_count,
        "total_matching": total_matching,
        "matching_indices": matching_indices,
        "page": page,
        "limit": limit,
    }


class ChartDataRequest(BaseModel):
    """Request body for chart-data endpoint."""
    chart_type: str  # bar, line, pie, scatter, histogram, area
    x_column: str
    y_column: str = ""
    aggregation: str = "sum"  # sum, avg, count, min, max
    group_by: str = ""
    null_handling: str = "exclude"  # exclude, category, zero
    histogram_bins: int = 10
    filters: dict | None = None  # Same format as filtered_dataset_post
    scatter_max_points: int = 5000  # Max points for scatter plot (downsampling)


def _apply_chart_filters(preview_list: list[dict], filters: dict | None) -> list[dict]:
    """Apply filters to data rows (same logic as filtered_dataset_post)."""
    if not filters:
        return preview_list

    substring_filters = {}
    selected_values_filters = {}

    for col, val in filters.items():
        if isinstance(val, dict) and "selected_values" in val:
            selected_vals = val["selected_values"]
            if isinstance(selected_vals, list) and len(selected_vals) > 0:
                normalized = set()
                for v in selected_vals:
                    if v is None:
                        normalized.add(None)
                    else:
                        normalized.add(str(v).strip())
                selected_values_filters[col] = normalized
        elif val is not None and str(val).strip():
            substring_filters[col] = str(val).strip().lower()

    filtered = []
    for row in preview_list:
        match = True

        for col, filter_val in substring_filters.items():
            cell_val = str(row.get(col, "")).lower()
            if filter_val not in cell_val:
                match = False
                break

        if not match:
            continue

        for col, allowed_values in selected_values_filters.items():
            cell_raw = row.get(col)
            if cell_raw is None:
                if None not in allowed_values:
                    match = False
                    break
            else:
                cell_str = str(cell_raw)
                if cell_str not in allowed_values:
                    match = False
                    break

        if match:
            filtered.append(row)

    return filtered


def _compute_chart_aggregation(
    data: list[dict],
    x_column: str,
    y_column: str,
    aggregation: str,
    group_by: str | None = None,
    null_handling: str = "exclude",
) -> dict:
    """Compute aggregated chart data from full dataset.

    Returns {"labels": [...], "datasets": [{"label": ..., "data": [...]}]}
    """
    if not data or not x_column:
        return {"labels": [], "datasets": [{"label": y_column or "Count", "data": []}]}

    # Filter nulls
    relevant_cols = [c for c in [x_column, y_column, group_by] if c]
    if null_handling == "exclude":
        data = [row for row in data if all(row.get(c) is not None and row.get(c) != "" for c in relevant_cols)]

    # Count without yColumn
    if aggregation == "count" and not y_column:
        counts: dict[str, int] = {}
        for row in data:
            key = str(row.get(x_column, "(null)"))
            counts[key] = counts.get(key, 0) + 1
        labels = sorted(counts.keys())
        return {"labels": labels, "datasets": [{"label": f"Count of {x_column}", "data": [counts[l] for l in labels]}]}

    def _get_y_value(row: dict, y_col: str, null_mode: str) -> float | None:
        """Extract y value from row, handling nulls based on null_mode."""
        if not y_col:
            return 1  # count mode with y_column
        raw = row.get(y_col)
        if raw is None or raw == "":
            if null_mode == "zero":
                return 0.0
            elif null_mode == "category":
                return None  # will be skipped
            else:  # exclude — should already be filtered, but safety fallback
                return None
        try:
            v = float(raw)
            if v != v:  # NaN check
                return None
            return v
        except (TypeError, ValueError):
            return None

    # Grouped aggregation
    if group_by:
        groups: dict[str, dict[str, list[float]]] = {}
        x_values: set[str] = set()

        for row in data:
            x = str(row.get(x_column, "(null)"))
            g = str(row.get(group_by, "(null)"))
            y = _get_y_value(row, y_column, null_handling)
            if y is None:
                continue
            x_values.add(x)
            if g not in groups:
                groups[g] = {}
            if x not in groups[g]:
                groups[g][x] = []
            groups[g][x].append(y)

        labels = sorted(x_values)
        datasets = []
        for group_name, x_data in groups.items():
            datasets.append({
                "label": group_name,
                "data": [_apply_agg(x_data.get(l, []), aggregation) for l in labels],
            })
        return {"labels": labels, "datasets": datasets}

    # Simple aggregation
    simple_groups: dict[str, list[float]] = {}
    for row in data:
        x = str(row.get(x_column, "(null)"))
        y = _get_y_value(row, y_column, null_handling)
        if y is None:
            continue
        if x not in simple_groups:
            simple_groups[x] = []
        simple_groups[x].append(y)

    labels = sorted(simple_groups.keys())
    return {
        "labels": labels,
        "datasets": [{"label": y_column or "Count", "data": [_apply_agg(simple_groups[l], aggregation) for l in labels]}],
    }


def _apply_agg(values: list[float], method: str) -> float:
    """Apply aggregation method to a list of numbers."""
    if not values:
        return 0
    if method == "sum":
        return sum(values)
    elif method == "avg":
        return sum(values) / len(values)
    elif method == "count":
        return len(values)
    elif method == "min":
        return min(values)
    elif method == "max":
        return max(values)
    return sum(values)


def _compute_chart_histogram(
    data: list[dict],
    column: str,
    num_bins: int = 10,
    null_handling: str = "exclude",
) -> dict:
    """Compute histogram bins from full dataset.

    Returns {"labels": [...], "data": [...]}
    """
    if not data or not column:
        return {"labels": [], "data": []}

    # Filter nulls
    if null_handling == "exclude":
        data = [row for row in data if row.get(column) is not None and row.get(column) != ""]

    values = []
    for row in data:
        try:
            v = float(row.get(column))
            if v == v:  # Not NaN
                values.append(v)
        except (TypeError, ValueError):
            continue

    if not values:
        return {"labels": [], "data": []}

    min_val = min(values)
    max_val = max(values)
    value_range = max_val - min_val or 1
    bin_width = value_range / num_bins

    bins = [0] * num_bins
    labels = []
    for i in range(num_bins):
        lo = min_val + i * bin_width
        hi = lo + bin_width
        labels.append(f"{lo:.1f}–{hi:.1f}")

    for v in values:
        idx = int((v - min_val) / bin_width)
        if idx >= num_bins:
            idx = num_bins - 1
        if idx < 0:
            idx = 0
        bins[idx] += 1

    return {"labels": labels, "data": bins}


def _compute_chart_scatter(
    data: list[dict],
    x_column: str,
    y_column: str,
    max_points: int = 5000,
    null_handling: str = "exclude",
) -> dict:
    """Compute scatter plot data from full dataset with optional downsampling.

    Returns {"datasets": [{"label": ..., "data": [{"x": ..., "y": ...}]}], "total_points": ..., "displayed_points": ...}
    """
    if not data or not x_column or not y_column:
        return {"datasets": [{"label": f"{x_column} vs {y_column}", "data": []}], "total_points": 0, "displayed_points": 0}

    # Filter nulls
    if null_handling == "exclude":
        data = [row for row in data if row.get(x_column) is not None and row.get(y_column) is not None and row.get(x_column) != "" and row.get(y_column) != ""]

    points = []
    for row in data:
        try:
            x = float(row.get(x_column))
            y = float(row.get(y_column))
            if x == x and y == y:  # Not NaN
                points.append({"x": x, "y": y})
        except (TypeError, ValueError):
            continue

    total_points = len(points)
    warning = None

    if total_points > max_points:
        # Downsample: take evenly spaced points
        step = total_points / max_points
        downsampled = [points[int(i * step)] for i in range(max_points)]
        points = downsampled
        warning = f"Showing {max_points} of {total_points} data points (downsampled)"

    return {
        "datasets": [{"label": f"{x_column} vs {y_column}", "data": points}],
        "total_points": total_points,
        "displayed_points": len(points),
        "warning": warning,
    }


@router.post("/{dataset_id}/chart-data")
async def get_chart_data(
    dataset_id: str,
    request: ChartDataRequest,
    current_user: User = Depends(get_current_active_user),
    session: AsyncSession = Depends(get_async_session),
):
    """Compute chart data from the full dataset.

    Supports aggregation (bar, line, pie, area), histogram, and scatter plots.
    Applies optional filters before computation.
    Returns chart-ready data (labels, datasets) for Chart.js.
    """
    result = await session.execute(select(Dataset).where(Dataset.id == dataset_id))
    dataset = result.scalar_one_or_none()

    if not dataset:
        raise HTTPException(status_code=404, detail="Dataset not found")

    project_result = await session.execute(
        select(Project).where(
            Project.id == dataset.project_id, Project.user_id == current_user.id
        )
    )
    if not project_result.scalar_one_or_none():
        raise HTTPException(status_code=403, detail="Access denied")

    # Get full data
    if not dataset.data_json or "data" not in dataset.data_json:
        return {
            "chart_type": request.chart_type,
            "labels": [],
            "datasets": [],
            "row_count": 0,
            "filtered_count": 0,
        }

    all_data = dataset.data_json["data"]
    row_count = len(all_data)

    # Apply filters if provided
    if request.filters:
        filtered_data = _apply_chart_filters(all_data, request.filters)
    else:
        filtered_data = all_data

    filtered_count = len(filtered_data)

    # Compute chart data based on type
    chart_type = request.chart_type
    warning = None

    if chart_type == "histogram":
        hist_result = _compute_chart_histogram(
            filtered_data,
            request.x_column,
            request.histogram_bins,
            request.null_handling,
        )
        return {
            "chart_type": chart_type,
            "labels": hist_result["labels"],
            "datasets": [{"label": f"Histogram of {request.x_column}", "data": hist_result["data"]}],
            "row_count": row_count,
            "filtered_count": filtered_count,
        }

    elif chart_type == "scatter":
        scatter_result = _compute_chart_scatter(
            filtered_data,
            request.x_column,
            request.y_column,
            request.scatter_max_points,
            request.null_handling,
        )
        response = {
            "chart_type": chart_type,
            "labels": [],
            "datasets": scatter_result["datasets"],
            "row_count": row_count,
            "filtered_count": filtered_count,
            "total_points": scatter_result.get("total_points", 0),
            "displayed_points": scatter_result.get("displayed_points", 0),
        }
        if scatter_result.get("warning"):
            response["warning"] = scatter_result["warning"]
        return response

    else:
        # Aggregated charts: bar, line, pie, area
        agg_result = _compute_chart_aggregation(
            filtered_data,
            request.x_column,
            request.y_column,
            request.aggregation,
            request.group_by or None,
            request.null_handling,
        )
        return {
            "chart_type": chart_type,
            "labels": agg_result["labels"],
            "datasets": agg_result["datasets"],
            "row_count": row_count,
            "filtered_count": filtered_count,
        }


@router.get("/{dataset_id}/unique-values")
async def get_unique_values(
    dataset_id: str,
    column: str = Query(..., description="Column name to get unique values for"),
    limit: int = Query(500, ge=1, le=5000, description="Max unique values to return"),
    current_user: User = Depends(get_current_active_user),
    session: AsyncSession = Depends(get_async_session),
):
    """Get unique values and their counts for a specific column.

    Returns value counts sorted by frequency (descending), useful for
    column multi-filter UI. Operates on the entire dataset.
    """
    result = await session.execute(select(Dataset).where(Dataset.id == dataset_id))
    dataset = result.scalar_one_or_none()

    if not dataset:
        raise HTTPException(status_code=404, detail="Dataset not found")

    project_result = await session.execute(
        select(Project).where(
            Project.id == dataset.project_id, Project.user_id == current_user.id
        )
    )
    if not project_result.scalar_one_or_none():
        raise HTTPException(status_code=403, detail="Access denied")

    # Verify column exists
    columns = _normalize_columns(dataset.columns)
    col_names = [c["name"] for c in columns]
    if column not in col_names:
        raise HTTPException(status_code=400, detail=f"Column '{column}' not found. Available: {col_names}")

    # Use data_json for computing unique values
    if not dataset.data_json or "data" not in dataset.data_json:
        return {"column": column, "values": [], "total_unique": 0}

    # Build a pandas Series from the column data for efficient value_counts
    all_data = dataset.data_json["data"]
    col_data = pd.Series([row.get(column) for row in all_data])

    # Compute value counts (exclude NaN/None by default, but include null count)
    null_count = int(col_data.isna().sum())
    non_null = col_data.dropna()

    # Convert all values to string for consistent comparison in the UI
    # (handles mixed types like int/float/string)
    if len(non_null) > 0:
        non_null_str = non_null.astype(str)
        value_counts = non_null_str.value_counts().head(limit)
        values_list = [
            {"value": str(v), "count": int(c)}
            for v, c in value_counts.items()
        ]
        total_unique = int(non_null_str.nunique())
    else:
        values_list = []
        total_unique = 0

    # Add null entry if there are nulls
    if null_count > 0:
        values_list.append({"value": None, "count": null_count})
        total_unique += 1  # Count null as a unique value

    return {
        "column": column,
        "values": values_list,
        "total_unique": total_unique,
    }


@router.get("/{dataset_id}/export")
async def export_dataset(
    dataset_id: str,
    format: str = "csv",
    columns: str | None = None,  # comma-separated column names
    limit: int = 0,  # 0 = all rows
    current_user: User = Depends(get_current_active_user),
    session: AsyncSession = Depends(get_async_session),
):
    """Export dataset to CSV, JSON, TSV, Excel, or Parquet. Returns a file download."""
    from fastapi.responses import StreamingResponse

    result = await session.execute(select(Dataset).where(Dataset.id == dataset_id))
    dataset = result.scalar_one_or_none()

    if not dataset:
        raise HTTPException(status_code=404, detail="Dataset not found")

    project_result = await session.execute(
        select(Project).where(
            Project.id == dataset.project_id, Project.user_id == current_user.id
        )
    )
    if not project_result.scalar_one_or_none():
        raise HTTPException(status_code=403, detail="Access denied")

    # Use data_json for export - generate DataFrame on-the-fly
    if not dataset.data_json or "data" not in dataset.data_json:
        raise HTTPException(status_code=400, detail="No data to export")

    df = pd.DataFrame(dataset.data_json["data"])

    # Filter columns if specified
    if columns:
        col_list = [c.strip() for c in columns.split(",") if c.strip()]
        valid_cols = [c for c in col_list if c in df.columns]
        if valid_cols:
            df = df[valid_cols]

    # Limit rows if specified
    if limit > 0:
        df = df.head(limit)

    safe_name = "".join(c if c.isalnum() or c in "-_ " else "_" for c in (dataset.name or "export"))

    if format == "csv":
        output = io.StringIO()
        df.to_csv(output, index=False)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{safe_name}.csv"'},
        )
    elif format == "json":
        content = df.to_json(orient="records", indent=2)
        return StreamingResponse(
            iter([content]),
            media_type="application/json",
            headers={"Content-Disposition": f'attachment; filename="{safe_name}.json"'},
        )
    elif format == "tsv":
        output = io.StringIO()
        df.to_csv(output, index=False, sep="\t")
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/tab-separated-values",
            headers={"Content-Disposition": f'attachment; filename="{safe_name}.tsv"'},
        )
    elif format == "excel":
        output = io.BytesIO()
        df.to_excel(output, index=False)
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{safe_name}.xlsx"'},
        )
    elif format == "parquet":
        output = io.BytesIO()
        df.to_parquet(output, index=False)
        output.seek(0)
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/octet-stream",
            headers={"Content-Disposition": f'attachment; filename="{safe_name}.parquet"'},
        )
    else:
        raise HTTPException(status_code=400, detail="Unsupported format. Use csv, json, tsv, excel, or parquet.")


@router.post("/{dataset_id}/export/db")
async def export_to_database(
    dataset_id: str,
    request: dict,
    current_user: User = Depends(get_current_active_user),
    session: AsyncSession = Depends(get_async_session),
):
    """Export dataset to a database table.

    Body: {
        "db_type": "postgresql" | "mysql" | "sqlite" | "mssql",
        "host": "localhost",
        "port": 5432,
        "database": "mydb",
        "username": "user",
        "password": "pass",
        "sslmode": "prefer",
        "table": "target_table",
        "mode": "create" | "append",
        "if_exists": "fail" | "replace" | "append"  (only for create mode)
    }
    """
    from sqlalchemy import create_engine, inspect, text

    # Validate dataset
    result = await session.execute(select(Dataset).where(Dataset.id == dataset_id))
    dataset = result.scalar_one_or_none()
    if not dataset:
        raise HTTPException(status_code=404, detail="Dataset not found")

    # Verify ownership
    project_result = await session.execute(
        select(Project).where(
            Project.id == dataset.project_id, Project.user_id == current_user.id
        )
    )
    if not project_result.scalar_one_or_none():
        raise HTTPException(status_code=403, detail="Access denied")

    if not dataset.data_json or "data" not in dataset.data_json:
        raise HTTPException(status_code=400, detail="No data to export")

    # Validate required fields
    table_name = request.get("table")
    if not table_name:
        raise HTTPException(status_code=400, detail="table is required")

    database = request.get("database")
    if not database:
        raise HTTPException(status_code=400, detail="database is required")

    db_type = request.get("db_type", "postgresql")
    mode = request.get("mode", "create")
    if_exists = request.get("if_exists", "fail")

    # Build connection URL
    try:
        url = _build_db_url(
            db_type,
            request.get("host", "localhost"),
            str(request.get("port", 5432)),
            database,
            request.get("username", ""),
            request.get("password", ""),
            sslmode=request.get("sslmode"),
        )
        engine = create_engine(url, connect_args=_get_connect_args(db_type))
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Connection failed: {str(e)}")

    # Load dataset into DataFrame
    df = pd.DataFrame(dataset.data_json["data"])

    # Convert datetime columns to strings for compatibility
    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            df[col] = df[col].astype(str)

    warnings = []

    try:
        if mode == "append":
            # Check if table exists
            inspector = inspect(engine)
            existing_tables = inspector.get_table_names()
            if table_name not in existing_tables:
                engine.dispose()
                raise HTTPException(
                    status_code=400,
                    detail=f"Table '{table_name}' does not exist. Use 'create' mode to create a new table.",
                )

            # Get existing table columns
            existing_columns = {col["name"] for col in inspector.get_columns(table_name)}
            df_columns = set(df.columns)

            # Check for schema mismatches
            extra_in_df = df_columns - existing_columns
            missing_in_df = existing_columns - df_columns

            if extra_in_df:
                warnings.append(
                    f"Columns in dataset but not in table will be dropped: {', '.join(sorted(extra_in_df))}"
                )
                df = df.drop(columns=list(extra_in_df))

            if missing_in_df:
                warnings.append(
                    f"Columns in table but not in dataset will be NULL: {', '.join(sorted(missing_in_df))}"
                )

        # Write to database
        df.to_sql(
            table_name,
            engine,
            if_exists=if_exists if mode == "create" else "append",
            index=False,
            chunksize=1000,
        )
        engine.dispose()

        return {
            "status": "success",
            "message": f"Exported {len(df)} rows to table '{table_name}'",
            "row_count": len(df),
            "table": table_name,
            "warnings": warnings if warnings else None,
        }

    except HTTPException:
        raise
    except Exception as e:
        engine.dispose()
        raise HTTPException(status_code=400, detail=f"Export failed: {str(e)}")


@router.get("")
async def list_datasets(
    project_id: str,
    page: int = 1,
    page_size: int = 20,
    current_user: User = Depends(get_current_active_user),
    session: AsyncSession = Depends(get_async_session),
):
    """List datasets for a project with pagination."""
    result = await session.execute(
        select(Project).where(
            Project.id == project_id, Project.user_id == current_user.id
        )
    )
    if not result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Project not found")

    # Count total
    from sqlalchemy import func
    count_result = await session.execute(
        select(func.count(Dataset.id)).where(Dataset.project_id == project_id)
    )
    total = count_result.scalar() or 0

    # Fetch page
    offset = (page - 1) * page_size
    datasets_result = await session.execute(
        select(Dataset)
        .where(Dataset.project_id == project_id)
        .order_by(Dataset.name)
        .offset(offset)
        .limit(page_size)
    )
    datasets = datasets_result.scalars().all()

    return {
        "datasets": [
            {
                "id": d.id, "name": d.name, "description": d.description,
                "file_name": d.file_name, "file_type": d.file_type,
                "row_count": d.row_count, "project_id": d.project_id,
            }
            for d in datasets
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@router.get("/{dataset_id}", response_model=DatasetResponse)
async def get_dataset(
    dataset_id: str,
    current_user: User = Depends(get_current_active_user),
    session: AsyncSession = Depends(get_async_session),
):
    """Get a specific dataset."""
    result = await session.execute(select(Dataset).where(Dataset.id == dataset_id))
    dataset = result.scalar_one_or_none()

    if not dataset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Dataset not found"
        )

    # Verify ownership
    project_result = await session.execute(
        select(Project).where(
            Project.id == dataset.project_id, Project.user_id == current_user.id
        )
    )
    if not project_result.scalar_one_or_none():
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN, detail="Access denied"
        )

    return dataset


@router.patch("/{dataset_id}")
async def update_dataset(
    dataset_id: str,
    request: dict,
    current_user: User = Depends(get_current_active_user),
    session: AsyncSession = Depends(get_async_session),
):
    """Update dataset name and/or description."""
    result = await session.execute(select(Dataset).where(Dataset.id == dataset_id))
    dataset = result.scalar_one_or_none()
    if not dataset:
        raise HTTPException(status_code=404, detail="Dataset not found")

    project_result = await session.execute(
        select(Project).where(
            Project.id == dataset.project_id, Project.user_id == current_user.id
        )
    )
    if not project_result.scalar_one_or_none():
        raise HTTPException(status_code=403, detail="Access denied")

    name = request.get("name")
    description = request.get("description")

    if name is not None:
        if not name.strip():
            raise HTTPException(status_code=400, detail="Name cannot be empty")
        dataset.name = name.strip()
    if description is not None:
        dataset.description = description.strip()

    await session.commit()
    await session.refresh(dataset)

    return {"status": "success", "message": "Dataset updated", "dataset": {"id": dataset.id, "name": dataset.name, "description": dataset.description}}


@router.delete("/{dataset_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_dataset(
    dataset_id: str,
    current_user: User = Depends(get_current_active_user),
    session: AsyncSession = Depends(get_async_session),
):
    """Delete a dataset."""
    result = await session.execute(select(Dataset).where(Dataset.id == dataset_id))
    dataset = result.scalar_one_or_none()

    if not dataset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Dataset not found"
        )

    # Verify ownership
    project_result = await session.execute(
        select(Project).where(
            Project.id == dataset.project_id, Project.user_id == current_user.id
        )
    )
    project = project_result.scalar_one_or_none()
    if not project:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN, detail="Access denied"
        )

    # Update project stats
    project.row_count -= dataset.row_count
    project.storage_bytes -= dataset.file_size

    await session.delete(dataset)
    await session.commit()

    return None


@router.post("/bulk-delete", response_model=dict)
async def bulk_delete_datasets(
    request: dict,
    current_user: User = Depends(get_current_active_user),
    session: AsyncSession = Depends(get_async_session),
):
    """Delete multiple datasets at once.

    Body: { dataset_ids: [...] }
    """
    dataset_ids = request.get("dataset_ids", [])
    
    if not dataset_ids:
        raise HTTPException(status_code=400, detail="dataset_ids is required")
    
    if len(dataset_ids) > 100:
        raise HTTPException(status_code=400, detail="Cannot delete more than 100 datasets at once")
    
    deleted_count = 0
    errors = []
    
    for dataset_id in dataset_ids:
        try:
            result = await session.execute(select(Dataset).where(Dataset.id == dataset_id))
            dataset = result.scalar_one_or_none()
            
            if not dataset:
                errors.append(f"Dataset {dataset_id} not found")
                continue
            
            # Verify ownership
            project_result = await session.execute(
                select(Project).where(
                    Project.id == dataset.project_id, Project.user_id == current_user.id
                )
            )
            project = project_result.scalar_one_or_none()
            if not project:
                errors.append(f"Access denied for dataset {dataset_id}")
                continue
            
            # Update project stats
            project.row_count -= dataset.row_count
            project.storage_bytes -= dataset.file_size
            
            await session.delete(dataset)
            deleted_count += 1
        except Exception as e:
            errors.append(f"Failed to delete {dataset_id}: {str(e)}")
    
    await session.commit()
    
    return {
        "status": "success",
        "deleted_count": deleted_count,
        "errors": errors,
    }


def _build_db_url(db_type: str, host: str, port: str, database: str, username: str, password: str, sslmode: str | None = None) -> str:
    """Build SQLAlchemy connection URL from parameters."""
    drivers = {
        "postgresql": "postgresql+psycopg2",
        "mysql": "mysql+pymysql",
        "sqlite": "sqlite",
        "oracle": "oracle+oracledb",
        "mssql": "mssql+pymssql",
    }
    driver = drivers.get(db_type)
    if not driver:
        raise HTTPException(status_code=400, detail=f"Unsupported database type: {db_type}")
    if db_type == "sqlite":
        return f"sqlite:///{database}"
    
    url = f"{driver}://{username}:{password}@{host}:{port}/{database}"
    if db_type == "postgresql" and sslmode:
        url += f"?sslmode={sslmode}"
    elif db_type == "mysql" and sslmode:
        url += f"?ssl-mode={sslmode}"
    return url


def _get_connect_args(db_type: str) -> dict:
    """Get connect_args for database connection. SQLite doesn't support connect_timeout."""
    if db_type in ("postgresql", "mysql"):
        return {"connect_timeout": 5}
    return {}


@router.post("/import/sqlite/upload")
async def upload_sqlite_file(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_active_user),
):
    """Upload a SQLite database file and return the temp path."""
    filename = file.filename or "database.db"
    file_ext = os.path.splitext(filename)[1].lower()
    
    # Validate SQLite file extension
    if file_ext not in (".db", ".sqlite", ".sqlite3"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File must be a SQLite database (.db, .sqlite, .sqlite3)"
        )
    
    content = await file.read()
    file_size = len(content)
    
    # Save to temp file with proper extension
    with tempfile.NamedTemporaryFile(delete=False, suffix=file_ext) as tmp:
        tmp.write(content)
        tmp_path = tmp.name
    
    return {
        "status": "success",
        "message": "SQLite file uploaded successfully",
        "path": tmp_path,
        "filename": filename,
        "file_size": file_size,
    }


@router.post("/import/db/test")
async def test_db_connection(request: dict):
    """Test a database connection."""
    from sqlalchemy import create_engine, text

    # Validate required fields
    database = request.get("database")
    if not database:
        raise HTTPException(status_code=400, detail="database is required")

    try:
        url = _build_db_url(
            request.get("db_type", "postgresql"),
            request.get("host", "localhost"),
            str(request.get("port", 5432)),
            database,
            request.get("username", ""),
            request.get("password", ""),
            sslmode=request.get("sslmode"),
        )
        engine = create_engine(url, connect_args=_get_connect_args(request.get("db_type", "postgresql")))
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        engine.dispose()
        return {"status": "success", "message": "Connection successful"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Connection failed: {str(e)}")


@router.post("/import/db/tables")
async def list_db_tables(request: dict):
    """List tables in a database."""
    from sqlalchemy import create_engine, inspect

    # Validate required fields
    database = request.get("database")
    if not database:
        raise HTTPException(status_code=400, detail="database is required")

    try:
        url = _build_db_url(
            request.get("db_type", "postgresql"),
            request.get("host", "localhost"),
            str(request.get("port", 5432)),
            database,
            request.get("username", ""),
            request.get("password", ""),
            sslmode=request.get("sslmode"),
        )
        engine = create_engine(url, connect_args=_get_connect_args(request.get("db_type", "postgresql")))
        inspector = inspect(engine)
        tables = inspector.get_table_names()
        engine.dispose()
        return {"status": "success", "tables": tables}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to list tables: {str(e)}")


@router.post("/import/db", response_model=dict)
async def import_from_database(
    request: dict,
    current_user: User = Depends(get_current_active_user),
    session: AsyncSession = Depends(get_async_session),
):
    """Import a database table as a dataset."""
    from sqlalchemy import create_engine, text

    # Validate required fields
    table_name = request.get("table")
    if not table_name:
        raise HTTPException(status_code=400, detail="table is required")

    database = request.get("database")
    if not database:
        raise HTTPException(status_code=400, detail="database is required")

    db_type = request.get("db_type", "postgresql")

    try:
        url = _build_db_url(
            db_type,
            request.get("host", "localhost"),
            str(request.get("port", 5432)),
            database,
            request.get("username", ""),
            request.get("password", ""),
            sslmode=request.get("sslmode"),
        )
        engine = create_engine(url, connect_args=_get_connect_args(db_type))
        
        # For SQLite, check if table exists before trying to read
        if db_type == "sqlite":
            from sqlalchemy import inspect
            inspector = inspect(engine)
            available_tables = inspector.get_table_names()
            if table_name not in available_tables:
                engine.dispose()
                raise HTTPException(
                    status_code=400,
                    detail=f"Table '{table_name}' not found in SQLite database. Available tables: {', '.join(available_tables) if available_tables else 'none'}"
                )
        
        df = pd.read_sql_table(table_name, engine)
        engine.dispose()

        # Convert datetime columns to ISO strings for JSON serialization
        for col in df.columns:
            if pd.api.types.is_datetime64_any_dtype(df[col]):
                df[col] = df[col].astype(str)

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read table: {str(e)}")

    if df.empty:
        raise HTTPException(status_code=400, detail="Table is empty")

    # Create dataset
    project_id = request.get("project_id")
    if not project_id:
        raise HTTPException(status_code=400, detail="project_id is required")

    project_result = await session.execute(
        select(Project).where(Project.id == project_id, Project.user_id == current_user.id)
    )
    if not project_result.scalar_one_or_none():
        raise HTTPException(status_code=403, detail="Access denied")

    # Get column info and preview
    try:
        columns = detect_columns(df)
        preview_data = get_preview_data(df)
        data_json = get_full_data_json(df)
        row_count = len(df)
    except Exception as e:
        logger.exception("Failed to process DataFrame for database import")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to process data: {str(e)}"
        )

    # Create dataset record
    try:
        dataset = Dataset(
            project_id=project_id,
            name=request.get("name", table_name),
            description=f"Imported from {db_type} table: {table_name}",
            file_name=f"{table_name}.csv",
            file_type="csv",
            row_count=row_count,
            columns=columns,
            data_json=data_json,
        )

        session.add(dataset)
        await session.commit()
        await session.refresh(dataset)

        return {
            "status": "success",
            "message": f"Imported {row_count} rows from '{table_name}'",
            "dataset_id": dataset.id,
            "row_count": row_count,
        }
    except Exception as e:
        await session.rollback()
        logger.exception("Failed to save dataset to database")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to save dataset: {str(e)}"
        )


@router.post("/merge", response_model=dict)
async def merge_datasets(
    request: dict,
    current_user: User = Depends(get_current_active_user),
    session: AsyncSession = Depends(get_async_session),
):
    """Merge (concatenate) multiple datasets into a new dataset.

    Body: { project_id, dataset_ids: [...], name: "...", strategy: "union" | "intersection" | "strict" }
    - union: keep all columns, fill missing with null (default)
    - intersection: keep only common columns
    - strict: fail if columns don't match exactly
    """
    project_id = request.get("project_id")
    dataset_ids = request.get("dataset_ids", [])
    name = request.get("name", "Merged Dataset")
    strategy = request.get("strategy", "union")

    if not project_id:
        raise HTTPException(status_code=400, detail="project_id is required")
    if len(dataset_ids) < 2:
        raise HTTPException(status_code=400, detail="At least 2 datasets required")

    # Verify project ownership
    project_result = await session.execute(
        select(Project).where(Project.id == project_id, Project.user_id == current_user.id)
    )
    if not project_result.scalar_one_or_none():
        raise HTTPException(status_code=403, detail="Access denied")

    # Fetch all datasets - use data_json for data access
    dfs = []
    all_columns = set()
    for ds_id in dataset_ids:
        result = await session.execute(select(Dataset).where(Dataset.id == ds_id, Dataset.project_id == project_id))
        ds = result.scalar_one_or_none()
        if not ds:
            raise HTTPException(status_code=404, detail=f"Dataset {ds_id} not found")
        if not ds.data_json or "data" not in ds.data_json:
            raise HTTPException(status_code=400, detail=f"Dataset '{ds.name}' has no data")
        df = pd.DataFrame(ds.data_json["data"])
        dfs.append(df)
        all_columns.update(df.columns)

    # Determine final columns based on strategy
    if strategy == "intersection":
        common_cols = set(dfs[0].columns)
        for df in dfs[1:]:
            common_cols &= set(df.columns)
        final_columns = sorted(common_cols)
    elif strategy == "strict":
        ref_cols = set(dfs[0].columns)
        mismatches = []
        for i, df in enumerate(dfs[1:], 2):
            ds_cols = set(df.columns)
            if ds_cols != ref_cols:
                missing = ref_cols - ds_cols
                extra = ds_cols - ref_cols
                detail = f"Dataset {i}"
                if missing:
                    detail += f" is missing: {sorted(missing)}"
                if extra:
                    detail += f" has extra: {sorted(extra)}"
                mismatches.append(detail)
        if mismatches:
            return {
                "status": "failed",
                "message": f"Column mismatch in strict mode: {'; '.join(mismatches)}",
                "details": mismatches,
            }
        final_columns = sorted(ref_cols)
    else:  # union
        final_columns = sorted(all_columns)

    if not final_columns:
        raise HTTPException(status_code=400, detail="No columns to merge")

    # Align all DataFrames to final columns
    aligned_dfs = []
    for df in dfs:
        for col in final_columns:
            if col not in df.columns:
                df[col] = None
        aligned_dfs.append(df[final_columns])

    # Concatenate
    try:
        merged_df = pd.concat(aligned_dfs, ignore_index=True)
    except Exception as e:
        logger.exception("Failed to concatenate DataFrames")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to merge datasets: {str(e)}"
        )

    # Get column info and preview
    try:
        columns = detect_columns(merged_df)
        preview_data = get_preview_data(merged_df)
        data_json = get_full_data_json(merged_df)
        row_count = len(merged_df)
    except Exception as e:
        logger.exception("Failed to process merged DataFrame")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to process merged data: {str(e)}"
        )

    # Create new dataset
    try:
        dataset = Dataset(
            project_id=project_id,
            name=name,
            description=f"Merged from {len(dataset_ids)} datasets ({strategy})",
            file_name=f"{name}.csv",
            file_type="csv",
            row_count=row_count,
            columns=columns,
            data_json=data_json,
        )
        session.add(dataset)
        await session.commit()
        await session.refresh(dataset)

        # Update project stats
        project_result2 = await session.execute(select(Project).where(Project.id == project_id))
        project = project_result2.scalar_one_or_none()
        if project:
            project.row_count = (project.row_count or 0) + row_count
            await session.commit()

        return {
            "status": "success",
            "message": f"Merged {len(dataset_ids)} datasets into '{name}' ({row_count} rows, {len(final_columns)} columns)",
            "dataset_id": dataset.id,
            "row_count": row_count,
            "column_count": len(final_columns),
        }
    except Exception as e:
        await session.rollback()
        logger.exception("Failed to save merged dataset to database")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to save merged dataset: {str(e)}"
        )


@router.post("/{dataset_id}/clone", response_model=DatasetResponse)
async def clone_dataset(
    dataset_id: str,
    request: dict | None = None,
    current_user: User = Depends(get_current_active_user),
    session: AsyncSession = Depends(get_async_session),
):
    """Clone a dataset with a new name.
    
    Body: { "name": "New Dataset Name" } (optional, defaults to "{original_name} (Copy)")
    """
    # Fetch the original dataset
    result = await session.execute(select(Dataset).where(Dataset.id == dataset_id))
    original = result.scalar_one_or_none()
    
    if not original:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Dataset not found"
        )
    
    # Verify ownership via project
    project_result = await session.execute(
        select(Project).where(
            Project.id == original.project_id, Project.user_id == current_user.id
        )
    )
    project = project_result.scalar_one_or_none()
    if not project:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN, detail="Access denied"
        )
    
    # Determine new name
    new_name = (request or {}).get("name")
    if not new_name:
        new_name = f"{original.name} (Copy)"
    
    # Create cloned dataset
    try:
        cloned = Dataset(
            name=new_name,
            description=original.description,
            project_id=original.project_id,
            file_name=original.file_name,
            file_size=original.file_size,
            file_type=original.file_type,
            columns=original.columns,
            data_json=original.data_json,
            row_count=original.row_count,
        )
        session.add(cloned)
        
        # Update project stats
        project.row_count = (project.row_count or 0) + original.row_count
        project.storage_bytes = (project.storage_bytes or 0) + original.file_size
        
        await session.commit()
        await session.refresh(cloned)
        
        return cloned
    except Exception as e:
        await session.rollback()
        logger.exception("Failed to clone dataset")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to clone dataset: {str(e)}"
        )


@router.post("/bulk-clone", response_model=dict)
async def bulk_clone_datasets(
    request: dict,
    current_user: User = Depends(get_current_active_user),
    session: AsyncSession = Depends(get_async_session),
):
    """Clone multiple datasets at once.
    
    Body: { dataset_ids: [...], name_prefix: "Copy of" }
    Each cloned dataset will be named "{name_prefix} {original_name}"
    """
    dataset_ids = request.get("dataset_ids", [])
    name_prefix = request.get("name_prefix", "Copy of")
    
    if not dataset_ids:
        raise HTTPException(status_code=400, detail="dataset_ids is required")
    
    if len(dataset_ids) > 100:
        raise HTTPException(status_code=400, detail="Cannot clone more than 100 datasets at once")
    
    cloned_count = 0
    errors = []
    
    for dataset_id in dataset_ids:
        try:
            result = await session.execute(select(Dataset).where(Dataset.id == dataset_id))
            original = result.scalar_one_or_none()
            
            if not original:
                errors.append(f"Dataset {dataset_id} not found")
                continue
            
            # Verify ownership via project
            project_result = await session.execute(
                select(Project).where(
                    Project.id == original.project_id, Project.user_id == current_user.id
                )
            )
            project = project_result.scalar_one_or_none()
            if not project:
                errors.append(f"Access denied for dataset {dataset_id}")
                continue
            
            # Create cloned dataset
            cloned = Dataset(
                name=f"{name_prefix} {original.name}",
                description=original.description,
                project_id=original.project_id,
                file_name=original.file_name,
                file_size=original.file_size,
                file_type=original.file_type,
                columns=original.columns,
                data_json=original.data_json,
                row_count=original.row_count,
            )
            session.add(cloned)
            
            # Update project stats
            project.row_count = (project.row_count or 0) + original.row_count
            project.storage_bytes = (project.storage_bytes or 0) + original.file_size
            
            cloned_count += 1
        except Exception as e:
            errors.append(f"Failed to clone {dataset_id}: {str(e)}")
    
    await session.commit()
    
    return {
        "status": "success",
        "cloned_count": cloned_count,
        "errors": errors,
    }


class CopyMoveRequest(BaseModel):
    dataset_ids: list[str]
    target_project_id: str
    action: str = "copy"  # "copy" or "move"


@router.post("/copy-move", response_model=dict)
async def copy_move_datasets(
    request: CopyMoveRequest,
    current_user: User = Depends(get_current_active_user),
    session: AsyncSession = Depends(get_async_session),
):
    """Copy or move datasets to another project.
    
    Body: { dataset_ids: [...], target_project_id: "...", action: "copy" | "move" }
    - copy: Clone datasets to target project (originals remain)
    - move: Transfer datasets to target project (originals are removed)
    """
    dataset_ids = request.dataset_ids
    target_project_id = request.target_project_id
    action = request.action
    
    if not dataset_ids:
        raise HTTPException(status_code=400, detail="dataset_ids is required")
    
    if len(dataset_ids) > 100:
        raise HTTPException(status_code=400, detail="Cannot process more than 100 datasets at once")
    
    if action not in ("copy", "move"):
        raise HTTPException(status_code=400, detail="action must be 'copy' or 'move'")
    
    # Verify target project ownership
    target_project_result = await session.execute(
        select(Project).where(
            Project.id == target_project_id, Project.user_id == current_user.id
        )
    )
    target_project = target_project_result.scalar_one_or_none()
    if not target_project:
        raise HTTPException(status_code=404, detail="Target project not found")
    
    processed_count = 0
    errors = []
    
    for dataset_id in dataset_ids:
        try:
            result = await session.execute(select(Dataset).where(Dataset.id == dataset_id))
            original = result.scalar_one_or_none()
            
            if not original:
                errors.append(f"Dataset {dataset_id} not found")
                continue
            
            # Verify ownership via project
            source_project_result = await session.execute(
                select(Project).where(
                    Project.id == original.project_id, Project.user_id == current_user.id
                )
            )
            source_project = source_project_result.scalar_one_or_none()
            if not source_project:
                errors.append(f"Access denied for dataset {dataset_id}")
                continue
            
            if action == "copy":
                # Create cloned dataset in target project
                cloned = Dataset(
                    name=original.name,
                    description=original.description,
                    project_id=target_project_id,
                    file_name=original.file_name,
                    file_size=original.file_size,
                    file_type=original.file_type,
                    columns=original.columns,
                    data_json=original.data_json,
                    row_count=original.row_count,
                )
                session.add(cloned)
                
                # Update target project stats
                target_project.row_count = (target_project.row_count or 0) + original.row_count
                target_project.storage_bytes = (target_project.storage_bytes or 0) + original.file_size
                
            else:  # move
                # Update source project stats
                source_project.row_count = (source_project.row_count or 0) - original.row_count
                source_project.storage_bytes = (source_project.storage_bytes or 0) - original.file_size
                
                # Update target project stats
                target_project.row_count = (target_project.row_count or 0) + original.row_count
                target_project.storage_bytes = (target_project.storage_bytes or 0) + original.file_size
                
                # Move dataset to target project
                original.project_id = target_project_id
            
            processed_count += 1
        except Exception as e:
            errors.append(f"Failed to process {dataset_id}: {str(e)}")
    
    await session.commit()
    
    return {
        "status": "success",
        "action": action,
        "processed_count": processed_count,
        "errors": errors,
    }
